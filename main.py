import asyncio, csv, io, json, os, shutil, sqlite3, uuid
import httpx

# Guard against PIL decompression-bomb attacks across the app
try:
    from PIL import Image as _PILImage
    _PILImage.MAX_IMAGE_PIXELS = int(os.getenv("MAX_IMAGE_PIXELS", "100000000"))  # 100 MP
except Exception:
    pass
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, BackgroundTasks, Depends, Request, Response
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import config
from pipeline.loader import load_folder, load_image, SUPPORTED
from pipeline.extractor import extract_batch
import database as db
from database import db as db_ctx
import vectorstore as vs
import chat
import memory
from auth import (
    bootstrap_super_admin,
    current_user,
    hash_secret,
    is_locked,
    make_token,
    normalize_identifier,
    record_login_attempt,
    require_role,
    verify_secret,
)
from pydantic import BaseModel

app = FastAPI(title="KONTACT — Catalog Vision RAG")

# ─── Rate limiting (slowapi) ─────────────────────────────────────────
RATE_LIMIT_ENABLED = (os.getenv("RATE_LIMIT_ENABLED") or "true").lower() == "true"
try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded
    limiter = Limiter(key_func=get_remote_address, enabled=RATE_LIMIT_ENABLED,
                      default_limits=["600/hour"])
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
except Exception as _e:
    print(f"[rate-limit] slowapi unavailable: {_e}")
    limiter = None

_cors_raw = (os.getenv("CORS_ORIGINS") or "http://localhost:8090").strip()
if _cors_raw == "*":
    CORS_ORIGINS = ["*"]
    CORS_CREDS = False
else:
    CORS_ORIGINS = [o.strip() for o in _cors_raw.split(",") if o.strip()]
    CORS_CREDS = True
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=CORS_CREDS,
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_BUILD = os.path.join(os.path.dirname(__file__), "frontend", "build")

SESSION_DAYS = int(os.getenv("SESSION_DAYS", "14"))
COOKIE_NAME = "kontact_session"
COOKIE_SECURE_ENV = (os.getenv("COOKIE_SECURE") or "auto").strip().lower()


def _set_session_cookie(response: Response, token: str, request: Request = None):
    if COOKIE_SECURE_ENV == "true":
        secure = True
    elif COOKIE_SECURE_ENV == "false":
        secure = False
    else:
        secure = bool(request and request.url.scheme == "https")
    response.set_cookie(
        key=COOKIE_NAME,
        value=token,
        max_age=SESSION_DAYS * 86400,
        httponly=True,
        samesite="lax",
        secure=secure,
        path="/",
    )


@app.on_event("startup")
def _bootstrap_auth():
    try:
        bootstrap_super_admin()
    except Exception as e:
        print(f"[auth] bootstrap error: {e}")
    # Tenancy: backfill legacy rows to super_admin (idempotent)
    try:
        db.run_tenancy_backfill()
    except Exception as e:
        print(f"[tenancy] backfill error: {e}")


from typing import Optional

class ChatRequest(BaseModel):
    question: str
    session_id: Optional[str] = None


# ─── AUTH ───

@app.post("/api/auth/login")
@(limiter.limit("10/minute") if limiter else (lambda f: f))
async def login(payload: dict, request: Request, response: Response):
    # Accept identifier or legacy 'email' field. Email + password only.
    identifier = (payload.get("identifier") or payload.get("email") or "").strip()
    secret = payload.get("secret") or payload.get("password") or ""
    ip = request.client.host if request.client else ""
    if not identifier or not secret:
        raise HTTPException(400, "email and password required")
    if is_locked(identifier):
        raise HTTPException(429, "too many attempts, locked")
    kind, value = normalize_identifier(identifier)
    if kind != "email":
        record_login_attempt(identifier, ip, 0)
        raise HTTPException(401, "invalid credentials")
    with db.db() as c:
        user = db.get_user_by_identifier(c, kind, value)
        if not user or not user["is_active"]:
            record_login_attempt(identifier, ip, 0)
            raise HTTPException(401, "invalid credentials")
        ok = False
        if user["password_hash"] and verify_secret(secret, user["password_hash"]):
            ok = True
        record_login_attempt(identifier, ip, 1 if ok else 0)
        if not ok:
            raise HTTPException(401, "invalid credentials")
        db.touch_login(c, user["uuid"])
    db.log_event_safe("login", "user", user["uuid"], user["uuid"], detail={"ip": ip})
    token = make_token(user["uuid"])
    _set_session_cookie(response, token, request)
    return {
        "uuid": user["uuid"],
        "name": user["name"],
        "email": user["email"],
        "phone_e164": user["phone_e164"],
        "role": user["role"],
        "token": token,
    }


@app.post("/api/auth/logout")
async def logout(response: Response, user=Depends(current_user)):
    response.delete_cookie(COOKIE_NAME, path="/")
    return {"ok": True}


@app.get("/api/auth/me")
async def me(user=Depends(current_user)):
    return {k: user[k] for k in ("uuid", "name", "email", "phone_e164", "role")}


@app.get("/api/users")
async def users_list(user=Depends(require_role("admin", "super_admin"))):
    with db.db() as c:
        rows = db.list_users(c)
    return [
        {k: r[k] for k in ("uuid", "name", "email", "phone_e164", "role", "is_active", "created_at", "last_login")}
        for r in rows
    ]


@app.post("/api/users")
async def users_create(payload: dict, user=Depends(require_role("admin", "super_admin"))):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name required")
    email = (payload.get("email") or "").strip().lower() or None
    phone_raw = (payload.get("phone") or "").strip() or None
    phone_e164 = None
    if phone_raw:
        kind, val = normalize_identifier(phone_raw)
        if kind != "phone":
            raise HTTPException(400, "invalid phone")
        phone_e164 = val
    if not email and not phone_e164:
        raise HTTPException(400, "email or phone required")
    password = payload.get("password") or ""
    pin = payload.get("pin") or ""
    if not password and not pin:
        raise HTTPException(400, "password or pin required")
    if pin and (not pin.isdigit() or len(pin) != 4):
        raise HTTPException(400, "pin must be 4 digits")
    role = payload.get("role", "user")
    if role not in ("admin", "user"):
        raise HTTPException(400, "invalid role")
    pw_hash = hash_secret(password) if password else None
    pin_hash = hash_secret(pin) if pin else None
    with db.db() as c:
        try:
            new_uuid = db.create_user(c, name, email, phone_e164, pw_hash, pin_hash, role, user["uuid"])
        except sqlite3.IntegrityError as e:
            raise HTTPException(409, f"duplicate: {e}")
    return {"uuid": new_uuid}


@app.patch("/api/users/{target_uuid}")
async def users_update(target_uuid: str, payload: dict, user=Depends(current_user)):
    is_admin = user["role"] in ("admin", "super_admin")
    is_self = user["uuid"] == target_uuid
    if not (is_admin or is_self):
        raise HTTPException(403, "forbidden")
    updates = {}
    for f in ("name", "email", "phone_e164"):
        if f in payload:
            updates[f] = payload[f]
    if "phone" in payload and payload["phone"]:
        kind, val = normalize_identifier(payload["phone"])
        if kind != "phone":
            raise HTTPException(400, "invalid phone")
        updates["phone_e164"] = val
    if "email" in updates and updates["email"]:
        updates["email"] = updates["email"].strip().lower()
    if "password" in payload and payload["password"]:
        updates["password_hash"] = hash_secret(payload["password"])
    if "pin" in payload and payload["pin"]:
        if not payload["pin"].isdigit() or len(payload["pin"]) != 4:
            raise HTTPException(400, "pin must be 4 digits")
        updates["pin_hash"] = hash_secret(payload["pin"])
    if is_admin:
        if "role" in payload and payload["role"] in ("admin", "user"):
            updates["role"] = payload["role"]
        if "is_active" in payload:
            updates["is_active"] = 1 if payload["is_active"] else 0
    with db.db() as c:
        row = db.get_user_by_uuid(c, target_uuid)
        if not row:
            raise HTTPException(404, "not found")
        if row["role"] == "super_admin" and "role" in updates:
            raise HTTPException(400, "cannot change super_admin role")
        try:
            db.update_user(c, target_uuid, updates)
        except sqlite3.IntegrityError as e:
            raise HTTPException(409, f"duplicate: {e}")
    return {"ok": True}


@app.delete("/api/users/{target_uuid}")
async def users_delete(target_uuid: str, user=Depends(require_role("admin", "super_admin"))):
    with db.db() as c:
        row = db.get_user_by_uuid(c, target_uuid)
        if not row:
            raise HTTPException(404, "not found")
        if row["role"] == "super_admin":
            raise HTTPException(400, "cannot delete super_admin")
        db.delete_user(c, target_uuid)
    return {"ok": True}


# ─── UPLOAD (just saves to queue, no processing) ───

@app.post("/api/upload")
@(limiter.limit("30/minute") if limiter else (lambda f: f))
async def upload_images(request: Request, files: list[UploadFile] = File(...), batch_id: str = Form(None), bg: BackgroundTasks = BackgroundTasks(), user=Depends(current_user)):
    if not batch_id:
        batch_id = str(uuid.uuid4())[:8]
    batch_dir = os.path.join(config.UPLOADS_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    # ── Read optional client_* form fields (browser-side geo + UA) ──
    form = await request.form()

    def _f(name):
        v = form.get(name)
        if v is None or v == "":
            return None
        try:
            fv = float(v)
            return fv if fv != 0 else None
        except (TypeError, ValueError):
            return None

    client_meta = {
        "client_gps_lat": _f("client_gps_lat"),
        "client_gps_lng": _f("client_gps_lng"),
        "client_gps_accuracy": _f("client_gps_accuracy"),
        "client_heading": _f("client_heading"),
        "client_timezone": form.get("client_timezone") or None,
        "client_timestamp": form.get("client_timestamp") or None,
        "client_user_agent": request.headers.get("user-agent"),
        "client_ip": request.client.host if request.client else None,
    }

    # Client-side EXIF sidecars + device signals
    import json as _json
    client_exif_raw = form.get("client_exif") or "[]"
    client_signals_raw = form.get("client_signals") or "{}"
    try:
        client_exif_list = _json.loads(client_exif_raw)
    except Exception:
        client_exif_list = []
    try:
        client_signals = _json.loads(client_signals_raw)
    except Exception:
        client_signals = {}

    # Build filename → sidecar map for per-file lookup
    exif_by_name = {x.get("filename"): x for x in client_exif_list if x.get("filename")}
    client_meta["exif_by_name"] = exif_by_name
    client_meta["device_signals"] = client_signals

    MAX_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", "104857600"))  # 100 MB default
    _ALLOWED_MIMES = {
        "image/jpeg", "image/png", "image/webp", "image/heic", "image/heif",
        "image/avif", "image/tiff", "image/gif", "application/pdf",
        "application/octet-stream",  # some browsers send this for HEIC
    }

    def _safe_filename(name: str) -> str:
        name = os.path.basename(name or "upload")
        # strip null bytes, traversal artifacts
        name = name.replace("\x00", "").replace("..", "_")
        return name[:200] or "upload"

    queued = []
    for f in files:
        safe_name = _safe_filename(f.filename)
        ext = os.path.splitext(safe_name)[1].lower()
        if ext not in SUPPORTED:
            continue
        # MIME sniff via Content-Type header (best-effort; PIL/fitz will reject malformed)
        ctype = (f.content_type or "").lower()
        if ctype and ctype not in _ALLOWED_MIMES:
            continue
        dest = os.path.join(batch_dir, safe_name)
        # Enforced size cap during streaming write
        written = 0
        try:
            with open(dest, "wb") as out:
                while True:
                    chunk = f.file.read(1 << 20)
                    if not chunk:
                        break
                    written += len(chunk)
                    if written > MAX_BYTES:
                        out.close()
                        try:
                            os.remove(dest)
                        except Exception:
                            pass
                        raise HTTPException(413, f"File {safe_name} exceeds {MAX_BYTES} bytes")
                    out.write(chunk)
        except HTTPException:
            raise
        except Exception as e:
            print(f"upload write error for {safe_name}: {e}")
            continue
        if ext == ".pdf":
            import fitz
            from PIL import Image as PILImage
            doc = fitz.open(dest)
            if len(doc) > 200:
                doc.close()
                try: os.remove(dest)
                except Exception: pass
                raise HTTPException(413, f"PDF {safe_name}: too many pages ({len(doc)} > 200)")
            base_name = os.path.splitext(safe_name)[0]
            for i, page in enumerate(doc, start=1):
                pix = page.get_pixmap(dpi=200)
                img = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
                page_name = f"{base_name}_page{i}.jpg"
                page_dest = os.path.join(batch_dir, page_name)
                img.save(page_dest, format="JPEG", quality=90)
                db.queue_add(batch_id, page_name, page_dest, owner_uuid=user["uuid"])
                queued.append(page_name)
            doc.close()
        else:
            db.queue_add(batch_id, safe_name, dest, owner_uuid=user["uuid"])
            queued.append(safe_name)

    # Auto-process in background after upload
    if queued:
        bg.add_task(_process_batch, batch_id, user["uuid"], client_meta)

    # Audit log (best-effort, per-file)
    for fname in queued:
        db.log_event_safe("upload", "document", None, user["uuid"],
                          detail={"batch_id": batch_id, "file": fname})

    return {"batch_id": batch_id, "queued": len(queued), "files": queued, "auto_processing": True}


@app.post("/api/upload/folder")
def upload_from_folder(folder_path: str, user=Depends(current_user)):
    if not os.path.isdir(folder_path):
        raise HTTPException(400, f"Not a folder: {folder_path}")
    batch_id = os.path.basename(folder_path.rstrip("/"))
    count = 0
    for f in sorted(os.listdir(folder_path)):
        ext = os.path.splitext(f)[1].lower()
        if ext in SUPPORTED:
            db.queue_add(batch_id, f, os.path.join(folder_path, f), owner_uuid=user["uuid"])
            count += 1
    if not count:
        raise HTTPException(400, "No supported images found")
    return {"batch_id": batch_id, "queued": count}


# ─── QUEUE STATUS ───

@app.get("/api/queue")
def queue_status(batch_id: str = None, user=Depends(current_user)):
    # Admins see global aggregate; users get filtered aggregate of own rows.
    if user["role"] in ("super_admin", "admin"):
        return db.queue_status(batch_id) if batch_id else db.queue_status()
    rows = db.get_queue_visible(user, batch_id)
    breakdown: dict = {}
    for r in rows:
        st = r.get("status") or "unknown"
        breakdown[st] = breakdown.get(st, 0) + 1
    return {"total": len(rows), **breakdown}


@app.get("/api/queue/batches")
def queue_batches(user=Depends(current_user)):
    return db.queue_batches()


@app.get("/api/queue/errors")
def queue_errors(batch_id: str = None, user=Depends(current_user)):
    rows = db.queue_errors(batch_id)
    if user["role"] in ("super_admin", "admin"):
        return rows
    uid = user["uuid"]
    return [r for r in rows if r.get("owner_uuid") == uid]


@app.post("/api/queue/retry/{queue_id}")
def retry_queue_item(queue_id: int, user=Depends(current_user)):
    # Edit-check: must own queue row OR be admin
    if user["role"] not in ("super_admin", "admin"):
        c = db._conn()
        try:
            row = c.execute("SELECT owner_uuid FROM queue WHERE id = ?", (queue_id,)).fetchone()
        finally:
            c.close()
        if not row:
            raise HTTPException(404, "Item not found")
        if row["owner_uuid"] != user["uuid"]:
            raise HTTPException(403, "forbidden")
    success = db.queue_retry(queue_id)
    if not success:
        raise HTTPException(404, "Item not found or not in error state")
    return {"retried": queue_id}


@app.delete("/api/batch/{batch_id}")
def delete_batch(batch_id: str, user=Depends(current_user)):
    # Non-admins: ensure they own every row in the batch
    if user["role"] not in ("super_admin", "admin"):
        c = db._conn()
        try:
            rows = c.execute(
                "SELECT DISTINCT owner_uuid FROM queue WHERE batch_id = ?", (batch_id,)
            ).fetchall()
        finally:
            c.close()
        owners = {r["owner_uuid"] for r in rows}
        if owners and owners - {user["uuid"]}:
            raise HTTPException(403, "forbidden")
    count = db.delete_batch(batch_id)
    if count == 0:
        raise HTTPException(404, "Batch not found")
    # Cascade to ChromaDB
    try:
        vs.delete_by_folder(batch_id)
    except Exception as e:
        print(f"[vs] delete_by_folder error: {e}")
    db.log_event_safe("delete", "batch", batch_id, user["uuid"], detail={"count": count})
    return {"deleted": batch_id, "count": count}


@app.get("/api/queue/pending")
def queue_pending(batch_id: str = None, user=Depends(current_user)):
    rows = db.queue_pending(batch_id)
    if user["role"] in ("super_admin", "admin"):
        return rows
    uid = user["uuid"]
    return [r for r in rows if r.get("owner_uuid") == uid]


# ─── PROCESS (runs extraction on pending queue items) ───

async def _process_batch(batch_id: str = None, owner_uuid: str = None, client_meta: dict = None):
    from pipeline.geocode import reverse_geocode
    from pipeline.imagequality import compute_phash, compute_blur, is_blurry as _is_blurry, find_near_dup

    pending = db.queue_pending(batch_id)
    if not pending:
        return {"processed": 0}

    images = []
    queue_map = {}
    for item in pending:
        try:
            img = load_image(item["file_path"])
            images.append(img)
            queue_map[img["file"]] = item
        except Exception as e:
            db.queue_update(item["id"], "error", error=str(e))

    if not images:
        return {"processed": 0, "errors": len(pending)}

    def on_progress(done, total, name):
        print(f"  [{done}/{total}] {name}")

    results = await extract_batch(images, on_progress=on_progress)

    cm = client_meta or {}
    done_count = 0
    for r in results:
        fname = r.get("source_file", "")
        item = queue_map.get(fname)
        if not item:
            continue
        if r.get("error"):
            db.queue_update(item["id"], "error", error=r["error"])
        else:
            db.queue_update(item["id"], "done", image_type=r.get("image_type"))
            folder = item["batch_id"]
            # Owner: explicit arg → queue row's owner_uuid → None
            row_owner = item.get("owner_uuid") if isinstance(item, dict) else None
            doc_owner = owner_uuid or row_owner

            # ── URL-QR enrichment: fetch profile pages, merge into contact ──
            try:
                from pipeline.url_resolver import resolve_profile_url
                qr_payloads_raw = r.get("qr_payloads")
                if qr_payloads_raw:
                    qrs = json.loads(qr_payloads_raw) if isinstance(qr_payloads_raw, str) else qr_payloads_raw
                    if isinstance(qrs, list):
                        for qr in qrs:
                            if not isinstance(qr, dict):
                                continue
                            if qr.get("type") != "url":
                                continue
                            url = qr.get("raw")
                            if not url:
                                continue
                            resolved = await resolve_profile_url(url)
                            if not resolved:
                                continue
                            contact = r.setdefault("contact", {})
                            for k_src, k_dst in (
                                ("name", "person"),
                                ("phone", "phone"),
                                ("phone_e164", "phone_e164"),
                                ("email", "email"),
                                ("company", "company"),
                                ("address", "address"),
                                ("job_title", "title"),
                            ):
                                if resolved.get(k_src) and not contact.get(k_dst):
                                    contact[k_dst] = resolved[k_src]
                            if not r.get("catalog_url"):
                                r["catalog_url"] = url
                            if resolved.get("company") and not r.get("company"):
                                r["company"] = resolved["company"]
                            r.setdefault("metadata", {})["url_resolution"] = {
                                "url": url,
                                "source_url": resolved.get("source_url"),
                                "social": resolved.get("social"),
                                "image_url": resolved.get("image_url"),
                            }
            except Exception as _e:
                print(f"[url_resolver] error for {fname}: {_e}")

            # ── Build metadata_extra (EXIF expansion + client + quality + geocode) ──
            src_path = r.get("source_path") or item.get("file_path") or ""
            exif_meta = r.get("metadata") or {}
            mx = {}

            # EXIF passthroughs
            for k in ("gps_altitude", "gps_heading", "gps_speed",
                      "lens_model", "focal_length", "f_number", "iso",
                      "exposure_time", "software", "sub_sec_time"):
                if exif_meta.get(k) is not None:
                    mx[k] = exif_meta[k]

            # GPS source decision
            exif_lat = exif_meta.get("gps_lat")
            exif_lng = exif_meta.get("gps_lng")
            client_lat = cm.get("client_gps_lat")
            client_lng = cm.get("client_gps_lng")
            final_lat, final_lng = None, None
            if exif_lat is not None and exif_lng is not None:
                final_lat, final_lng = exif_lat, exif_lng
                mx["gps_source"] = "exif"
            elif client_lat is not None and client_lng is not None:
                final_lat, final_lng = client_lat, client_lng
                mx["gps_lat"] = client_lat
                mx["gps_lng"] = client_lng
                mx["gps_source"] = "browser"
                if cm.get("client_gps_accuracy") is not None:
                    mx["gps_accuracy"] = cm["client_gps_accuracy"]
                if cm.get("client_heading") is not None and mx.get("gps_heading") is None:
                    mx["gps_heading"] = cm["client_heading"]

            # Reverse geocode (cached + rate-limited)
            if final_lat is not None and final_lng is not None:
                try:
                    geo = await reverse_geocode(final_lat, final_lng)
                    if geo:
                        mx["country"] = geo.get("country")
                        mx["city"] = geo.get("city")
                        mx["address_full"] = geo.get("address_full")
                except Exception:
                    pass

            # Client fields
            if cm.get("client_timezone"):
                mx["client_timezone"] = cm["client_timezone"]
            if cm.get("client_user_agent"):
                mx["client_user_agent"] = cm["client_user_agent"]
            if cm.get("client_ip"):
                mx["client_ip"] = cm["client_ip"]
            if cm.get("client_timestamp"):
                mx["client_timestamp"] = cm["client_timestamp"]

            # Client-side EXIF sidecar (fills gaps when server-side EXIF is stripped/missing)
            sidecar = cm.get("exif_by_name", {}).get(os.path.basename(src_path))
            if sidecar:
                for k in ("gps_lat", "gps_lng", "gps_altitude", "gps_heading", "gps_speed",
                          "date_taken", "camera_make", "camera_model", "lens_model",
                          "focal_length", "f_number", "iso", "exposure_time", "software",
                          "img_width", "img_height"):
                    if not exif_meta.get(k) and sidecar.get(k) is not None:
                        exif_meta[k] = sidecar[k]
                        # Also surface in metadata_extra so flat columns are populated
                        if k in ("gps_altitude", "gps_heading", "gps_speed",
                                 "lens_model", "focal_length", "f_number", "iso",
                                 "exposure_time", "software"):
                            mx[k] = sidecar[k]
                if not mx.get("gps_source") and sidecar.get("gps_lat"):
                    mx["gps_source"] = "exif"  # came from client-side EXIF read
                    if final_lat is None:
                        mx["gps_lat"] = sidecar["gps_lat"]
                        mx["gps_lng"] = sidecar["gps_lng"]

            # Device signals → JSON blob
            import json as _json_ds
            mx["device_signals"] = _json_ds.dumps(cm.get("device_signals") or {})

            # Image quality + dedup
            if src_path and os.path.isfile(src_path):
                ph = compute_phash(src_path)
                if ph:
                    mx["image_phash"] = ph
                    conn = db._conn()
                    try:
                        dup = find_near_dup(conn, ph)
                    finally:
                        conn.close()
                    if dup:
                        mx["near_dup_of"] = dup
                bs = compute_blur(src_path)
                if bs is not None:
                    mx["blur_score"] = bs
                    mx["is_blurry"] = _is_blurry(bs)

            # ── Source channel inference ───────────────────────────
            folder_lc = (folder or "").lower()
            if folder_lc.startswith("wechat_") or folder_lc.startswith("wechat-"):
                _src_ch = "wechat"
            elif folder_lc.startswith("email_") or folder_lc.startswith("email-"):
                _src_ch = "email"
            elif (r.get("metadata") or {}).get("source") == "url_ingest" or folder_lc.startswith("url_"):
                _src_ch = "url"
            else:
                _src_ch = "upload"
            mx["source_channel"] = _src_ch

            db.insert_extraction(folder, r, owner_uuid=doc_owner, is_shared=0, metadata_extra=mx)
            vs.index_record(folder, r)
            done_count += 1

    # Save JSON
    bid = batch_id or "batch"
    out_path = os.path.join(config.EXTRACTIONS_DIR, f"{bid}.json")
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    return {"processed": done_count, "errors": len(results) - done_count}


@app.post("/api/process")
async def process_queue(batch_id: str = None, user=Depends(current_user)):
    result = await _process_batch(batch_id)
    return result


@app.post("/api/process/background")
async def process_background(batch_id: str = None, bg: BackgroundTasks = None, user=Depends(current_user)):
    pending = db.queue_pending(batch_id)
    if not pending:
        return {"status": "complete", "pending": 0, "batch_id": batch_id, "message": "Nothing pending"}
    bg.add_task(_process_batch, batch_id)
    return {"status": "processing", "pending": len(pending), "batch_id": batch_id}


# ─── SEARCH ───

@app.get("/api/search")
def search(q: str = Query(...), limit: int = 20, user=Depends(current_user)):
    rows = db.search(q, limit)
    # Post-filter by visibility (admins see all)
    if user["role"] in ("super_admin", "admin"):
        return rows
    uid = user["uuid"]
    return [r for r in rows if r.get("owner_uuid") == uid]


@app.get("/api/search/semantic")
def semantic_search(q: str = Query(...), n: int = 5, user=Depends(current_user)):
    return vs.query(q, n)


# ─── DATA ───

@app.get("/api/data")
def get_data(folder: str = None, limit: int = 500, user=Depends(current_user)):
    return db.get_documents_visible(user, folder=folder, limit=limit)


@app.get("/api/documents/{doc_id}")
def get_document_one(doc_id: int, user=Depends(current_user)):
    doc = db.get_document_visible(doc_id, user)
    if not doc:
        raise HTTPException(404, "document not found")
    return doc


@app.get("/api/stats")
def stats(user=Depends(current_user)):
    return db.get_stats_visible(user)


@app.post("/api/index")
def index_all(user=Depends(current_user)):
    db_count = db.load_all_extractions()
    vs_count = vs.index_all_from_json()
    return {"db_records": db_count, "vector_records": vs_count}


# ─── CHAT ───

@app.post("/api/chat")
@(limiter.limit("60/minute") if limiter else (lambda f: f))
async def chat_endpoint(req: ChatRequest, request: Request, user=Depends(current_user)):
    session_id = req.session_id or str(uuid.uuid4())[:8]
    result = await chat.ask(req.question, session_id=session_id, user=user)
    return result


@app.get("/api/chat/sessions")
def chat_sessions(user=Depends(current_user)):
    return db.list_sessions(user=user)


@app.get("/api/chat/history/{session_id}")
def chat_history(session_id: str, user=Depends(current_user)):
    # Non-admin: ensure they own the session (or it is unowned legacy)
    if user["role"] not in ("super_admin", "admin"):
        owner = db.session_owner(session_id)
        if owner and owner != user["uuid"]:
            raise HTTPException(403, "forbidden")
    return db.get_chat_history(session_id, user=user)


@app.delete("/api/chat/sessions/{session_id}")
def delete_chat_session(session_id: str, user=Depends(current_user)):
    if user["role"] not in ("super_admin", "admin"):
        owner = db.session_owner(session_id)
        if owner and owner != user["uuid"]:
            raise HTTPException(403, "forbidden")
    db.delete_session(session_id)
    return {"deleted": session_id}


# ─── EXPORT ───

@app.get("/api/export/json")
def export_json(user=Depends(current_user)):
    data = db.export_all_visible(user)
    return JSONResponse(content=data, headers={"Content-Disposition": "attachment; filename=kontact_export.json"})


@app.get("/api/contacts")
def get_contacts(user=Depends(current_user)):
    return db.get_contacts_table_visible(user)


@app.get("/api/contacts/{uuid}")
def get_contact_one(uuid: str, user=Depends(current_user)):
    ct = db.get_contact_visible(uuid, user)
    if not ct:
        raise HTTPException(404, "contact not found")
    return ct


@app.get("/api/products")
def get_products(user=Depends(current_user)):
    return db.get_products_table_visible(user)


@app.get("/api/products/{uuid}")
def get_product_one(uuid: str, user=Depends(current_user)):
    p = db.get_product_visible(uuid, user)
    if not p:
        raise HTTPException(404, "product not found")
    return p


@app.get("/api/documents/metadata")
def get_documents_metadata(limit: int = 500, user=Depends(current_user)):
    return db.get_documents_with_metadata(limit)


@app.post("/api/migrate")
def run_migration(user=Depends(require_role("admin", "super_admin"))):
    result = db.populate_normalized_tables()
    return {"status": "ok", **result}


@app.get("/api/dashboard")
def dashboard(user=Depends(current_user)):
    return db.get_dashboard_stats()


@app.get("/api/export/xlsx")
def export_xlsx(user=Depends(current_user)):
    from openpyxl import Workbook
    data = db.export_all_visible(user)
    if not data:
        raise HTTPException(404, "No data")

    wb = Workbook()

    # Sheet 1: Products
    ws_prod = wb.active
    ws_prod.title = "Products"
    prod_headers = ["company", "product_name", "model", "specs", "category", "price", "folder", "source_file"]
    ws_prod.append(prod_headers)
    for row in data:
        prods = row.get("products", [])
        if isinstance(prods, str):
            try:
                prods = json.loads(prods)
            except Exception:
                prods = []
        if not isinstance(prods, list):
            continue
        for p in prods:
            if not isinstance(p, dict):
                continue
            ws_prod.append([
                p.get("company", "") or row.get("company", ""),
                p.get("product_name", "") or p.get("name", ""),
                p.get("model", ""),
                p.get("specs", "") if isinstance(p.get("specs"), str) else json.dumps(p.get("specs", ""), ensure_ascii=False),
                p.get("category", ""),
                p.get("price", ""),
                row.get("folder", ""),
                row.get("source_file", ""),
            ])

    # Sheet 2: Contacts (filtered by visibility — uses normalized contacts table)
    ws_cont = wb.create_sheet("Contacts")
    cont_headers = ["company", "person", "phone", "email", "website", "address", "folder"]
    ws_cont.append(cont_headers)
    contacts = db.get_all_contact_records_visible(user)
    for ct in contacts:
        ws_cont.append([ct.get(h, "") for h in cont_headers])

    # Sheet 3: Summary
    ws_sum = wb.create_sheet("Summary")
    sum_headers = ["company", "document_count", "product_count", "has_contact"]
    ws_sum.append(sum_headers)
    company_info = {}
    contact_companies = {ct["company"] for ct in contacts if ct.get("company")}
    for row in data:
        comp = row.get("company", "") or "Unknown"
        if comp not in company_info:
            company_info[comp] = {"doc_count": 0, "prod_count": 0}
        company_info[comp]["doc_count"] += 1
        prods = row.get("products", [])
        if isinstance(prods, str):
            try:
                prods = json.loads(prods)
            except Exception:
                prods = []
        if isinstance(prods, list):
            company_info[comp]["prod_count"] += len(prods)
    for comp, info in sorted(company_info.items()):
        ws_sum.append([comp, info["doc_count"], info["prod_count"], "Yes" if comp in contact_companies else "No"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kontact_export.xlsx"},
    )


@app.get("/api/export/csv")
def export_csv(user=Depends(current_user)):
    data = db.export_all_visible(user)
    if not data:
        raise HTTPException(404, "No data")
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["folder", "source_file", "image_type", "company", "title", "raw_text"])
    writer.writeheader()
    for row in data:
        writer.writerow({k: row.get(k, "") for k in writer.fieldnames})
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=kontact_export.csv"})


# ─── IMAGE SERVING (for citations) ───

def _safe_under(base_dir: str, candidate: str) -> bool:
    try:
        base_real = os.path.realpath(base_dir)
        cand_real = os.path.realpath(candidate)
        # Ensure cand stays within base_real (avoids symlink/.. escapes)
        return cand_real == base_real or cand_real.startswith(base_real + os.sep)
    except Exception:
        return False


@app.get("/api/image/{folder}/{filename:path}")
async def serve_image(folder: str, filename: str, user=Depends(current_user)):
    from fastapi.responses import FileResponse
    if '..' in folder or '..' in filename or folder.startswith('/') or filename.startswith('/'):
        raise HTTPException(400, "Invalid path")
    if '\x00' in folder or '\x00' in filename:
        raise HTTPException(400, "Invalid path")

    is_admin = user.get("role") in ("super_admin", "admin")
    try:
        c = db._conn()
        try:
            # Find document owning this image; non-admin must own it
            if is_admin:
                doc_row = c.execute(
                    "SELECT source_path, owner_uuid FROM documents WHERE folder = ? AND source_file = ?",
                    (folder, filename),
                ).fetchone()
            else:
                doc_row = c.execute(
                    "SELECT source_path, owner_uuid FROM documents WHERE folder = ? AND source_file = ? AND owner_uuid = ?",
                    (folder, filename, user["uuid"]),
                ).fetchone()

            # If no document row yet (still in queue), check queue ownership
            if not doc_row:
                if is_admin:
                    q_row = c.execute(
                        "SELECT file_path, owner_uuid FROM queue WHERE batch_id = ? AND file_name = ?",
                        (folder, filename),
                    ).fetchone()
                else:
                    q_row = c.execute(
                        "SELECT file_path, owner_uuid FROM queue WHERE batch_id = ? AND file_name = ? AND owner_uuid = ?",
                        (folder, filename, user["uuid"]),
                    ).fetchone()
                if q_row and q_row["file_path"] and os.path.isfile(q_row["file_path"]):
                    fp = q_row["file_path"]
                    if not _safe_under(config.UPLOADS_DIR, fp):
                        raise HTTPException(403, "outside uploads dir")
                    return FileResponse(fp)
                raise HTTPException(404, "not found")

            # Use authoritative source_path from document, fall back to constructed path
            fp = doc_row["source_path"] or os.path.join(config.UPLOADS_DIR, folder, filename)
            if not _safe_under(config.UPLOADS_DIR, fp):
                raise HTTPException(403, "outside uploads dir")
            if os.path.isfile(fp):
                return FileResponse(fp)
        finally:
            c.close()
    except HTTPException:
        raise
    except Exception as e:
        print(f"Image serve error: {e} for {folder}/{filename}")
    raise HTTPException(404, "Image not found")


# ─── STREAMING CHAT ───

@app.post("/api/chat/stream")
async def chat_stream(req: ChatRequest, user=Depends(current_user)):
    from starlette.responses import StreamingResponse as SSEResponse
    from tools import execute_tool
    session_id = req.session_id or str(uuid.uuid4())[:8]

    async def event_generator():
        # Send session info
        yield f"event: session\ndata: {json.dumps({'session_id': session_id})}\n\n"

        # Build context with learning memory
        yield f"event: status\ndata: {json.dumps({'step': 'Searching knowledge base...'})}\n\n"
        context = chat._build_context(req.question)
        system_prompt = chat._build_system_prompt(user=user)

        yield f"event: status\ndata: {json.dumps({'step': 'Querying LLM...'})}\n\n"

        # Get history
        history = db.get_chat_history(session_id, limit=6, user=user) if session_id else []

        messages = [{"role": "system", "content": system_prompt}]
        for h in history[-6:]:
            messages.append({"role": h["role"], "content": h["content"]})
        messages.append({"role": "user", "content": f"Context:\n{context}\n\nQuestion: {req.question}"})

        _stream_headers = {
            "Authorization": f"Bearer {config.OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        }

        async def _do_stream(msgs):
            """Async generator: stream LLM call, yielding content deltas."""
            payload = {
                "model": config.VISION_MODEL,
                "messages": msgs,
                "max_tokens": 2000,
                "temperature": 0.1,
                "stream": True,
            }
            async with httpx.AsyncClient() as client:
                async with client.stream("POST", config.OPENROUTER_BASE,
                                         json=payload, headers=_stream_headers, timeout=120) as r:
                    async for line in r.aiter_lines():
                        if not line.startswith("data: "):
                            continue
                        data_str = line[6:]
                        if data_str.strip() == "[DONE]":
                            break
                        try:
                            chunk = json.loads(data_str)
                            delta = chunk.get("choices", [{}])[0].get("delta", {}).get("content", "")
                            if delta:
                                yield delta
                        except Exception:
                            continue

        # First LLM call — buffer it (don't stream yet, might contain tool calls)
        full_answer = ""
        async for delta in _do_stream(messages):
            full_answer += delta

        # Tool execution loop (up to MAX_TOOL_ITERATIONS)
        for _tool_iter in range(chat.MAX_TOOL_ITERATIONS):
            if not chat._has_tool_calls(full_answer):
                break

            import time as _time
            tool_calls = chat._parse_tool_calls(full_answer)
            tool_results = []
            for tc in tool_calls:
                tname = tc["name"]
                yield f"event: status\ndata: {json.dumps({'step': 'Running ' + tname + '...'})}\n\n"
                t0 = _time.time()
                r = execute_tool(tname, tc["args"], user=user)
                dur = round(_time.time() - t0, 2)
                step_done = '✓ ' + tname + ' (' + str(dur) + 's)'
                yield f"event: status\ndata: {json.dumps({'step': step_done})}\n\n"
                tool_results.append(f"[RESULT: {tname}]\n{r}\n[/RESULT]")

            result_block = "\n\n".join(tool_results)
            messages.append({"role": "assistant", "content": full_answer})
            messages.append({"role": "user", "content": (
                f"Tool results:\n\n{result_block}\n\n"
                "Now provide your final answer. Do NOT include [TOOL:] markers. "
                "Just give the clean answer with data from the tool results."
            )})

            yield f"event: status\ndata: {json.dumps({'step': 'Generating answer...'})}\n\n"

            # Get final answer — buffer again in case of more tool calls
            full_answer = ""
            async for delta in _do_stream(messages):
                full_answer += delta

        # Strip any leftover tool markers
        full_answer = chat._TOOL_PATTERN.sub("", full_answer).strip()

        # Now stream the clean final answer to the client
        yield f"event: content\ndata: {json.dumps({'text': full_answer})}\n\n"

        # Get sources
        sources = []
        for s in vs.query(req.question, n_results=3):
            src_file = s["metadata"]["source_file"]
            folder = s["metadata"]["folder"]
            sources.append({
                "folder": folder,
                "file": src_file,
                "company": s["metadata"].get("company", ""),
                "image_url": f"/api/image/{folder}/{src_file}",
            })

        # Save to history
        db.save_chat(session_id, "user", req.question, user_uuid=user["uuid"])
        db.save_chat(session_id, "assistant", full_answer, user_uuid=user["uuid"])

        yield f"event: done\ndata: {json.dumps({'sources': sources, 'session_id': session_id})}\n\n"

    return SSEResponse(event_generator(), media_type="text/event-stream")


# ─── FEEDBACK & MEMORY ───

class FeedbackRequest(BaseModel):
    session_id: str
    question: str
    answer: str
    rating: str  # "up" or "down"


@app.post("/api/feedback")
def submit_feedback(req: FeedbackRequest, user=Depends(current_user)):
    memory.save_feedback(req.session_id, req.question, req.answer, req.rating)
    return {"saved": True, "rating": req.rating}


@app.get("/api/memories")
def list_memories(limit: int = 10, user=Depends(current_user)):
    return memory.get_memories(limit=limit)


@app.get("/api/config")
def get_config(user=Depends(current_user)):
    return {
        "vision_model": config.VISION_MODEL,
        "embedding_model": config.EMBEDDING_MODEL,
        "provider": "OpenRouter",
    }


@app.get("/health")
def health():
    return {"status": "ok"}


# ─── SYNC CHANNELS (WeChat watcher, IMAP email, URL ingest, vCard, manual edit) ───

from sync import wechat_watcher, email_poll, vcard as vcard_mod, snapshot as snapshot_mod
from pipeline.loader import extract_qr_codes


class WeChatStartRequest(BaseModel):
    folder_path: str


class WeChatFolderScanRequest(BaseModel):
    folder_path: str
    vendor_company: Optional[str] = None


class WeChatChatAssign(BaseModel):
    vendor_company: Optional[str] = None
    contact_uuid: Optional[str] = None
    notes: Optional[str] = None


class EmailStartRequest(BaseModel):
    host: str
    user: str
    app_password: str
    mailbox: Optional[str] = "INBOX"
    interval_minutes: Optional[int] = 5


class URLIngestRequest(BaseModel):
    url: str
    vendor_company: Optional[str] = None


class ContactMergeRequest(BaseModel):
    keep_uuid: str
    merge_uuid: str


@app.post("/api/sync/wechat/start")
def sync_wechat_start(req: WeChatStartRequest, user=Depends(current_user)):
    try:
        return wechat_watcher.start_watcher(req.folder_path)
    except FileNotFoundError as e:
        raise HTTPException(400, str(e))
    except RuntimeError as e:
        raise HTTPException(500, str(e))


@app.post("/api/sync/wechat/stop")
def sync_wechat_stop(user=Depends(current_user)):
    return wechat_watcher.stop_watcher()


@app.get("/api/sync/wechat/status")
def sync_wechat_status(user=Depends(current_user)):
    return wechat_watcher.watcher_status()


@app.post("/api/sync/wechat-folder")
def sync_wechat_folder(req: WeChatFolderScanRequest, user=Depends(current_user)):
    try:
        return wechat_watcher.scan_folder_once(req.folder_path, req.vendor_company)
    except FileNotFoundError as e:
        raise HTTPException(400, str(e))


@app.get("/api/sync/wechat/chats")
def sync_wechat_chats(user=Depends(current_user)):
    return db.wechat_map_list()


@app.post("/api/sync/wechat/chats/{chat_hash}/assign")
def sync_wechat_chats_assign(chat_hash: str, req: WeChatChatAssign, user=Depends(current_user)):
    return db.wechat_map_upsert(chat_hash, req.vendor_company, req.contact_uuid, req.notes)


@app.delete("/api/sync/wechat/chats/{chat_hash}")
def sync_wechat_chats_delete(chat_hash: str, user=Depends(current_user)):
    n = db.wechat_map_delete(chat_hash)
    if not n:
        raise HTTPException(404, "chat_hash not found")
    return {"deleted": chat_hash}


@app.post("/api/sync/email/start")
def sync_email_start(req: EmailStartRequest, user=Depends(current_user)):
    try:
        return email_poll.start_poller(
            req.host, req.user, req.app_password,
            interval_minutes=req.interval_minutes or 5,
            mailbox=req.mailbox or "INBOX",
        )
    except RuntimeError as e:
        raise HTTPException(500, str(e))


@app.post("/api/sync/email/stop")
def sync_email_stop(user=Depends(current_user)):
    return email_poll.stop_poller()


@app.get("/api/sync/email/status")
def sync_email_status(user=Depends(current_user)):
    return email_poll.poller_status()


@app.post("/api/ingest/url")
def ingest_url(req: URLIngestRequest, user=Depends(current_user)):
    import hashlib as _hl, time as _time
    slug = "".join(ch if ch.isalnum() else "_" for ch in (req.vendor_company or "url"))[:30] or "url"
    batch_id = f"url_{slug}_{int(_time.time())}"
    out_dir = os.path.join(config.UPLOADS_DIR, batch_id)
    try:
        path = snapshot_mod.snapshot_url(req.url, out_dir)
    except httpx.HTTPError as e:
        raise HTTPException(400, f"Fetch failed: {e}")
    fname = os.path.basename(path)
    ext = os.path.splitext(fname)[1].lower()
    if ext not in SUPPORTED:
        # Saved but not queueable for vision extractor (e.g. HTML)
        return {"saved": path, "queued": False, "batch_id": batch_id,
                "reason": f"extension {ext} not in SUPPORTED (saved only)"}
    try:
        with open(path, "rb") as f:
            fhash = _hl.sha256(f.read()).hexdigest()
    except Exception:
        fhash = None
    db.queue_add(batch_id, fname, path,
                source_channel="url", source_sender=req.url, file_hash=fhash,
                owner_uuid=user["uuid"])
    return {"saved": path, "queued": True, "batch_id": batch_id}


@app.get("/api/contacts/{uuid}/vcard")
def contact_vcard(uuid: str, user=Depends(current_user)):
    ct = db.get_contact_visible(uuid, user)
    if not ct:
        raise HTTPException(404, "contact not found")
    body = vcard_mod.to_vcard(ct)
    fname = (ct.get("person") or ct.get("company") or uuid).replace(" ", "_") + ".vcf"
    return StreamingResponse(
        io.BytesIO(body.encode("utf-8")),
        media_type="text/vcard",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.get("/api/export/vcards.zip")
def export_vcards_zip(user=Depends(current_user)):
    contacts = db.get_all_contact_records_visible(user)
    if not contacts:
        raise HTTPException(404, "no contacts")
    blob = vcard_mod.bulk_vcards_zip(contacts)
    return StreamingResponse(
        io.BytesIO(blob),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=kontact_vcards.zip"},
    )


@app.patch("/api/contacts/{uuid}")
def patch_contact(uuid: str, body: dict, user=Depends(current_user)):
    ct = db.get_contact_by_uuid(uuid)
    if not ct:
        raise HTTPException(404, "contact not found")
    if not db.can_edit(ct, user):
        raise HTTPException(403, "forbidden")
    n = db.update_contact(uuid, body or {})
    db.log_event_safe("edit", "contact", uuid, user["uuid"], detail={"fields": list((body or {}).keys())})
    return {"updated": n, "uuid": uuid}


@app.patch("/api/products/{uuid}")
def patch_product(uuid: str, body: dict, user=Depends(current_user)):
    p = db.get_product_by_uuid(uuid)
    if not p:
        raise HTTPException(404, "product not found")
    if not db.can_edit(p, user):
        raise HTTPException(403, "forbidden")
    n = db.update_product(uuid, body or {})
    if not n:
        raise HTTPException(400, "no editable fields")
    return {"updated": n, "uuid": uuid}


@app.post("/api/contacts/merge")
def contacts_merge(req: ContactMergeRequest, user=Depends(current_user)):
    keep = db.get_contact_by_uuid(req.keep_uuid)
    drop = db.get_contact_by_uuid(req.merge_uuid)
    if not keep or not drop:
        raise HTTPException(404, "contact not found")
    if not (db.can_edit(keep, user) and db.can_edit(drop, user)):
        raise HTTPException(403, "forbidden — both contacts must be editable by you")
    result = db.merge_contacts(req.keep_uuid, req.merge_uuid)
    if not result.get("merged"):
        raise HTTPException(400, result.get("reason", "merge failed"))
    db.log_event_safe("merge", "contact", req.keep_uuid, user["uuid"],
                      detail={"merged_uuid": req.merge_uuid})
    return result


@app.delete("/api/contacts/{uuid}")
def delete_contact_one(uuid: str, user=Depends(current_user)):
    ct = db.get_contact_by_uuid(uuid)
    if not ct:
        raise HTTPException(404, "contact not found")
    if not db.can_edit(ct, user):
        raise HTTPException(403, "forbidden")
    db.delete_contact(uuid)
    db.log_event_safe("delete", "contact", uuid, user["uuid"])
    return {"deleted": uuid}


@app.delete("/api/products/{uuid}")
def delete_product_one(uuid: str, user=Depends(current_user)):
    p = db.get_product_by_uuid(uuid)
    if not p:
        raise HTTPException(404, "product not found")
    if not db.can_edit(p, user):
        raise HTTPException(403, "forbidden")
    db.delete_product(uuid)
    db.log_event_safe("delete", "product", uuid, user["uuid"])
    return {"deleted": uuid}


# ─── Share toggles ───

class ShareRequest(BaseModel):
    is_shared: bool


@app.post("/api/documents/{doc_id}/share")
def share_document(doc_id: int, req: ShareRequest, user=Depends(current_user)):
    doc = db.get_document(doc_id)
    if not doc:
        raise HTTPException(404, "document not found")
    if not db.can_edit(doc, user):
        raise HTTPException(403, "forbidden")
    db.set_document_shared(doc_id, req.is_shared)
    db.log_event_safe("share", "document", str(doc_id), user["uuid"], detail={"is_shared": req.is_shared})
    return {"doc_id": doc_id, "is_shared": req.is_shared}


@app.post("/api/contacts/{uuid}/share")
def share_contact(uuid: str, req: ShareRequest, user=Depends(current_user)):
    ct = db.get_contact_by_uuid(uuid)
    if not ct:
        raise HTTPException(404, "contact not found")
    if not db.can_edit(ct, user):
        raise HTTPException(403, "forbidden")
    db.set_contact_shared(uuid, req.is_shared)
    db.log_event_safe("share", "contact", uuid, user["uuid"], detail={"is_shared": req.is_shared})
    return {"uuid": uuid, "is_shared": req.is_shared}


@app.post("/api/products/{uuid}/share")
def share_product(uuid: str, req: ShareRequest, user=Depends(current_user)):
    p = db.get_product_by_uuid(uuid)
    if not p:
        raise HTTPException(404, "product not found")
    if not db.can_edit(p, user):
        raise HTTPException(403, "forbidden")
    db.set_product_shared(uuid, req.is_shared)
    db.log_event_safe("share", "product", uuid, user["uuid"], detail={"is_shared": req.is_shared})
    return {"uuid": uuid, "is_shared": req.is_shared}


@app.post("/api/documents/{doc_id}/rescan-qr")
def rescan_qr(doc_id: int, user=Depends(current_user)):
    doc = db.get_document(doc_id)
    if not doc:
        raise HTTPException(404, "document not found")
    if not db.can_edit(doc, user):
        raise HTTPException(403, "forbidden")
    src = doc.get("source_path")
    if not src or not os.path.isfile(src):
        # try uploads layout fallback
        candidate = os.path.join(config.UPLOADS_DIR, doc.get("folder") or "", doc.get("source_file") or "")
        if os.path.isfile(candidate):
            src = candidate
        else:
            raise HTTPException(400, "source image not on disk")
    try:
        qrs = extract_qr_codes(src)
    except Exception as e:
        raise HTTPException(500, f"QR decode failed: {e}")
    db.update_document_qr(doc_id, qrs)

    # Merge any parsed messenger/vcard data into contacts for this document
    merged_into = []
    try:
        for q in qrs or []:
            parsed = q.get("parsed") or {}
            qtype = q.get("type")
            update_fields = {}
            if qtype in ("vcard", "mecard"):
                for k in ("person", "company", "phone", "email", "website", "address"):
                    if parsed.get(k):
                        update_fields[k] = parsed[k]
            elif qtype in ("whatsapp", "viber", "telegram", "line", "signal", "wechat"):
                key_map = {
                    "whatsapp": "whatsapp", "viber": "viber", "telegram": "telegram",
                    "line": "line_id", "signal": "signal_phone", "wechat": "wechat_id",
                }
                handle = parsed.get("handle") or parsed.get("phone")
                if handle:
                    update_fields[key_map[qtype]] = handle
            if not update_fields:
                continue
            # Apply to contacts attached to this document
            c = db._conn()
            try:
                rows = c.execute("SELECT uuid FROM contacts WHERE document_id = ?", (doc_id,)).fetchall()
            finally:
                c.close()
            for r in rows:
                db.update_contact(r["uuid"], update_fields)
                merged_into.append(r["uuid"])
    except Exception:
        pass

    return {"doc_id": doc_id, "qr_count": len(qrs), "qrs": qrs, "merged_into_contacts": merged_into}


# ─── STARTUP: auto-start watcher if WECHAT_WATCH_DIR set ───
@app.on_event("startup")
def _auto_start_sync():
    watch = os.getenv("WECHAT_WATCH_DIR", "").strip()
    if watch and os.path.isdir(watch):
        try:
            wechat_watcher.start_watcher(watch)
            print(f"[sync] WeChat watcher auto-started on {watch}")
        except Exception as e:
            print(f"[sync] auto-start failed: {e}")


@app.on_event("startup")
def _start_vector_pruner():
    """Daily prune of orphan ChromaDB vectors (folder no longer in documents)."""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except Exception:
        return

    def _prune():
        try:
            with db_ctx() as c:
                folders = {r["folder"] for r in c.execute("SELECT DISTINCT folder FROM documents").fetchall()}
            n = vs.prune_orphans(folders)
            if n:
                print(f"[vs] pruned {n} orphan vectors")
        except Exception as e:
            print(f"[vs] prune error: {e}")

    sched = BackgroundScheduler()
    sched.add_job(_prune, "interval", hours=24, id="vs_prune")
    sched.start()
    print("[vs] vector pruner scheduled every 24h")


# ─── AGGREGATIONS ───

def _vis_where(user, alias=""):
    """Build a WHERE-fragment string + params using visibility_clause."""
    vc, vp = db.visibility_clause(alias, user)
    return (f" AND {vc}" if vc else "", vp)


@app.get("/api/aggregations/locations")
def agg_locations(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = "WHERE gps_lat IS NOT NULL"
    if vc:
        where += f" AND {vc}"
    c = db._conn()
    try:
        rows = c.execute(
            f"""SELECT country, city,
                       COUNT(*) AS doc_count,
                       AVG(gps_lat) AS lat_center,
                       AVG(gps_lng) AS lng_center
                FROM documents
                {where}
                GROUP BY country, city
                ORDER BY doc_count DESC""",
            vp,
        ).fetchall()
    finally:
        c.close()
    return [dict(r) for r in rows]


@app.get("/api/aggregations/countries")
def agg_countries(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where_docs = f"WHERE {vc}" if vc else ""
    c = db._conn()
    try:
        doc_rows = c.execute(
            f"""SELECT COALESCE(country, '') AS country, COUNT(*) AS doc_count
                FROM documents {where_docs}
                GROUP BY country""",
            vp,
        ).fetchall()
        contact_rows = c.execute(
            f"""SELECT COALESCE(d.country, '') AS country, COUNT(*) AS contact_count
                FROM contacts ct
                LEFT JOIN documents d ON d.uuid = ct.document_uuid
                {('WHERE ' + vc.replace('owner_uuid', 'ct.owner_uuid').replace('is_shared', 'ct.is_shared')) if vc else ''}
                GROUP BY country""",
            vp,
        ).fetchall()
        product_rows = c.execute(
            f"""SELECT COALESCE(d.country, '') AS country, COUNT(*) AS product_count
                FROM products p
                LEFT JOIN documents d ON d.uuid = p.document_uuid
                {('WHERE ' + vc.replace('owner_uuid', 'p.owner_uuid').replace('is_shared', 'p.is_shared')) if vc else ''}
                GROUP BY country""",
            vp,
        ).fetchall()
    finally:
        c.close()
    by_country = {}
    for r in doc_rows:
        by_country.setdefault(r["country"], {"country": r["country"], "doc_count": 0,
                                              "contact_count": 0, "product_count": 0})
        by_country[r["country"]]["doc_count"] = r["doc_count"]
    for r in contact_rows:
        by_country.setdefault(r["country"], {"country": r["country"], "doc_count": 0,
                                              "contact_count": 0, "product_count": 0})
        by_country[r["country"]]["contact_count"] = r["contact_count"]
    for r in product_rows:
        by_country.setdefault(r["country"], {"country": r["country"], "doc_count": 0,
                                              "contact_count": 0, "product_count": 0})
        by_country[r["country"]]["product_count"] = r["product_count"]
    return sorted(by_country.values(), key=lambda x: x["doc_count"], reverse=True)


@app.get("/api/aggregations/timeline")
def agg_timeline(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = "WHERE date_taken IS NOT NULL AND date_taken != ''"
    if vc:
        where += f" AND {vc}"
    c = db._conn()
    try:
        rows = c.execute(
            f"""SELECT substr(REPLACE(date_taken, ':', '-'), 1, 10) AS date,
                       COUNT(*) AS doc_count,
                       GROUP_CONCAT(DISTINCT folder) AS batches
                FROM documents
                {where}
                GROUP BY date
                ORDER BY date DESC""",
            vp,
        ).fetchall()
    finally:
        c.close()
    out = []
    for r in rows:
        d = dict(r)
        batches_raw = d.get("batches") or ""
        d["batches"] = [b for b in batches_raw.split(",") if b]
        out.append(d)
    return out


@app.get("/api/aggregations/cameras")
def agg_cameras(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = "WHERE (camera_make IS NOT NULL OR camera_model IS NOT NULL)"
    if vc:
        where += f" AND {vc}"
    c = db._conn()
    try:
        rows = c.execute(
            f"""SELECT COALESCE(camera_make, '') AS camera_make,
                       COALESCE(camera_model, '') AS camera_model,
                       COUNT(*) AS doc_count
                FROM documents {where}
                GROUP BY camera_make, camera_model
                ORDER BY doc_count DESC""",
            vp,
        ).fetchall()
    finally:
        c.close()
    return [dict(r) for r in rows]


@app.get("/api/aggregations/messengers")
def agg_messengers(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = f"WHERE {vc}" if vc else ""
    c = db._conn()
    try:
        rows = c.execute(
            f"""SELECT uuid, person, company, messengers,
                       whatsapp, wechat_id, wechat_qr_url, viber, telegram,
                       line_id, signal_phone
                FROM contacts {where}""",
            vp,
        ).fetchall()
    finally:
        c.close()
    out = {"whatsapp": [], "wechat": [], "viber": [], "line": [], "telegram": [], "signal": []}
    for r in rows:
        meta = {
            "contact_uuid": r["uuid"],
            "person": r["person"] or "",
            "company": r["company"] or "",
        }
        # JSON messengers blob
        raw = r["messengers"]
        if raw:
            try:
                entries = json.loads(raw) if isinstance(raw, str) else raw
            except (json.JSONDecodeError, TypeError):
                entries = None
            if isinstance(entries, list):
                for e in entries:
                    if not isinstance(e, dict):
                        continue
                    plat = (e.get("platform") or "").lower()
                    if plat in out:
                        out[plat].append({**meta, "handle": e.get("handle") or e.get("phone") or ""})
        # Flat platform columns
        if r["whatsapp"]:
            out["whatsapp"].append({**meta, "handle": r["whatsapp"]})
        if r["wechat_qr_url"] or r["wechat_id"]:
            out["wechat"].append({**meta, "handle": r["wechat_id"] or r["wechat_qr_url"]})
        if r["viber"]:
            out["viber"].append({**meta, "handle": r["viber"]})
        if r["telegram"]:
            out["telegram"].append({**meta, "handle": r["telegram"]})
        if r["line_id"]:
            out["line"].append({**meta, "handle": r["line_id"]})
        if r["signal_phone"]:
            out["signal"].append({**meta, "handle": r["signal_phone"]})
    return out


@app.get("/api/aggregations/qr-codes")
def agg_qr_codes(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = "WHERE qr_payloads IS NOT NULL AND qr_payloads != ''"
    if vc:
        where += f" AND {vc}"
    c = db._conn()
    try:
        rows = c.execute(
            f"SELECT uuid, source_file, qr_payloads FROM documents {where}",
            vp,
        ).fetchall()
    finally:
        c.close()
    out = []
    for r in rows:
        try:
            qrs = json.loads(r["qr_payloads"]) if isinstance(r["qr_payloads"], str) else r["qr_payloads"]
        except Exception:
            continue
        if not isinstance(qrs, list):
            continue
        for q in qrs:
            if not isinstance(q, dict):
                continue
            out.append({
                "document_uuid": r["uuid"],
                "source_file": r["source_file"],
                "qr_type": q.get("type"),
                "qr_raw": q.get("raw") or q.get("data"),
                "qr_parsed": q.get("parsed"),
            })
    return out


@app.get("/api/aggregations/quality")
def agg_quality(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = f"WHERE {vc}" if vc else ""
    c = db._conn()
    try:
        blurry = c.execute(
            f"SELECT COUNT(*) FROM documents {where + (' AND ' if where else 'WHERE ')}is_blurry = 1",
            vp,
        ).fetchone()[0]
        low_res = c.execute(
            f"SELECT COUNT(*) FROM documents {where + (' AND ' if where else 'WHERE ')}"
            "(img_width IS NOT NULL AND img_width < 800)",
            vp,
        ).fetchone()[0]
        near_dups = c.execute(
            f"SELECT COUNT(*) FROM documents {where + (' AND ' if where else 'WHERE ')}near_dup_of IS NOT NULL",
            vp,
        ).fetchone()[0]
        items = c.execute(
            f"""SELECT uuid, source_file, blur_score, img_width, img_height, near_dup_of, is_blurry
                FROM documents
                {where + (' AND ' if where else 'WHERE ')}
                (is_blurry = 1 OR near_dup_of IS NOT NULL OR (img_width IS NOT NULL AND img_width < 800))
                ORDER BY blur_score DESC LIMIT 200""",
            vp,
        ).fetchall()
    finally:
        c.close()
    return {
        "blurry": blurry,
        "low_res": low_res,
        "near_dups": near_dups,
        "items": [dict(r) for r in items],
    }


@app.get("/api/aggregations/duplicates")
def agg_duplicates(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = "WHERE near_dup_of IS NOT NULL"
    if vc:
        where += f" AND {vc}"
    c = db._conn()
    try:
        rows = c.execute(
            f"""SELECT near_dup_of AS cluster_id, uuid, source_file, image_phash, blur_score
                FROM documents
                {where}
                ORDER BY near_dup_of""",
            vp,
        ).fetchall()
    finally:
        c.close()
    clusters = {}
    for r in rows:
        cid = r["cluster_id"]
        clusters.setdefault(cid, {"cluster_id": cid, "members": []})
        clusters[cid]["members"].append({
            "uuid": r["uuid"],
            "source_file": r["source_file"],
            "phash": r["image_phash"],
            "blur_score": r["blur_score"],
        })
    # Include the original doc each cluster points to
    if clusters:
        ids = list(clusters.keys())
        placeholders = ",".join("?" for _ in ids)
        c = db._conn()
        try:
            origs = c.execute(
                f"SELECT uuid, source_file, image_phash, blur_score FROM documents "
                f"WHERE uuid IN ({placeholders})",
                ids,
            ).fetchall()
        finally:
            c.close()
        orig_map = {r["uuid"]: dict(r) for r in origs}
        for cid, cluster in clusters.items():
            if cid in orig_map:
                o = orig_map[cid]
                cluster["members"].insert(0, {
                    "uuid": o["uuid"],
                    "source_file": o["source_file"],
                    "phash": o["image_phash"],
                    "blur_score": o["blur_score"],
                })
    return list(clusters.values())


@app.get("/api/aggregations/sync-sources")
def agg_sync_sources(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = f"WHERE {vc}" if vc else ""
    c = db._conn()
    try:
        rows = c.execute(
            f"SELECT folder, source_channel, COUNT(*) AS cnt FROM documents {where} "
            "GROUP BY folder, source_channel",
            vp,
        ).fetchall()
    finally:
        c.close()
    buckets = {"upload": 0, "wechat": 0, "email": 0, "url": 0}
    for r in rows:
        folder = (r["folder"] or "").lower()
        ch = (r["source_channel"] or "").lower()
        if ch in ("wechat", "wechat_desktop") or folder.startswith("wechat"):
            buckets["wechat"] += r["cnt"]
        elif ch == "email" or folder.startswith("email"):
            buckets["email"] += r["cnt"]
        elif ch == "url" or folder.startswith("url"):
            buckets["url"] += r["cnt"]
        else:
            buckets["upload"] += r["cnt"]
    return [{"source": k, "doc_count": v} for k, v in buckets.items()]


@app.get("/api/aggregations/pricing")
def agg_pricing(user=Depends(current_user)):
    import re as _re
    vc, vp = db.visibility_clause("", user)
    where = f"WHERE {vc}" if vc else ""
    c = db._conn()
    try:
        rows = c.execute(
            f"SELECT company, category, price FROM products {where}",
            vp,
        ).fetchall()
    finally:
        c.close()
    groups = {}
    for r in rows:
        company = r["company"] or ""
        category = r["category"] or ""
        price_raw = r["price"] or ""
        nums = _re.findall(r"\d+(?:\.\d+)?", str(price_raw))
        if not nums:
            continue
        try:
            val = float(nums[0])
        except Exception:
            continue
        key = (company, category)
        groups.setdefault(key, []).append(val)
    out = []
    for (company, category), vals in groups.items():
        vals_sorted = sorted(vals)
        n = len(vals_sorted)
        median = vals_sorted[n // 2] if n % 2 else (vals_sorted[n // 2 - 1] + vals_sorted[n // 2]) / 2
        out.append({
            "company": company,
            "category": category,
            "min_price": min(vals_sorted),
            "max_price": max(vals_sorted),
            "median_price": median,
            "count": n,
        })
    out.sort(key=lambda x: x["count"], reverse=True)
    return out


@app.get("/api/aggregations/map-points")
def agg_map_points(user=Depends(current_user)):
    vc, vp = db.visibility_clause("", user)
    where = "WHERE gps_lat IS NOT NULL AND gps_lng IS NOT NULL"
    if vc:
        where += f" AND {vc}"
    c = db._conn()
    try:
        rows = c.execute(
            f"""SELECT uuid AS doc_uuid, gps_lat AS lat, gps_lng AS lng,
                       city, country, source_file, company
                FROM documents
                {where}
                ORDER BY date_taken DESC LIMIT 5000""",
            vp,
        ).fetchall()
    finally:
        c.close()
    return [dict(r) for r in rows]


# ─── TAGS CRUD ───

class TagCreateReq(BaseModel):
    name: str
    color: Optional[str] = None


class TagAttachReq(BaseModel):
    tag_uuid: str


@app.get("/api/tags")
def tags_list(user=Depends(current_user)):
    with db.db() as c:
        return db.list_tags(c, user)


@app.post("/api/tags")
def tags_create(req: TagCreateReq, user=Depends(current_user)):
    with db.db() as c:
        tag_uuid = db.create_tag(c, req.name, req.color, user["uuid"])
    return {"uuid": tag_uuid, "name": req.name}


@app.delete("/api/tags/{tag_uuid}")
def tags_delete(tag_uuid: str, user=Depends(current_user)):
    with db.db() as c:
        tag = db.get_tag(c, tag_uuid)
        if not tag:
            raise HTTPException(404, "tag not found")
        is_admin = user["role"] in ("super_admin", "admin")
        if not is_admin and tag.get("created_by") != user["uuid"]:
            raise HTTPException(403, "forbidden")
        db.delete_tag(c, tag_uuid)
    return {"deleted": tag_uuid}


@app.post("/api/documents/{doc_uuid}/tags")
def doc_tag_attach(doc_uuid: str, req: TagAttachReq, user=Depends(current_user)):
    with db.db() as c:
        db.tag_document(c, doc_uuid, req.tag_uuid, user["uuid"])
    return {"document_uuid": doc_uuid, "tag_uuid": req.tag_uuid}


@app.delete("/api/documents/{doc_uuid}/tags/{tag_uuid}")
def doc_tag_detach(doc_uuid: str, tag_uuid: str, user=Depends(current_user)):
    with db.db() as c:
        n = db.untag_document(c, doc_uuid, tag_uuid)
    if not n:
        raise HTTPException(404, "tag attachment not found")
    return {"deleted": True}


@app.get("/api/documents/{doc_uuid}/tags")
def doc_tags_list(doc_uuid: str, user=Depends(current_user)):
    with db.db() as c:
        return db.list_doc_tags(c, doc_uuid)


# ─── NOTES CRUD ───

class NoteCreateReq(BaseModel):
    entity_type: str
    entity_uuid: str
    body: str


class NoteUpdateReq(BaseModel):
    body: str


@app.get("/api/notes")
def notes_list(entity_type: str = Query(...), entity_uuid: str = Query(...),
               user=Depends(current_user)):
    with db.db() as c:
        return db.list_notes(c, entity_type, entity_uuid, user)


@app.post("/api/notes")
def notes_create(req: NoteCreateReq, user=Depends(current_user)):
    with db.db() as c:
        nuuid = db.create_note(c, req.entity_type, req.entity_uuid, req.body, user["uuid"])
    return {"uuid": nuuid}


@app.patch("/api/notes/{note_uuid}")
def notes_update(note_uuid: str, req: NoteUpdateReq, user=Depends(current_user)):
    with db.db() as c:
        n = db.update_note(c, note_uuid, req.body, user["uuid"])
    if n == -1:
        raise HTTPException(403, "forbidden")
    if n == 0:
        raise HTTPException(404, "note not found")
    return {"updated": n}


@app.delete("/api/notes/{note_uuid}")
def notes_delete(note_uuid: str, user=Depends(current_user)):
    is_admin = user["role"] in ("super_admin", "admin")
    with db.db() as c:
        n = db.delete_note(c, note_uuid, user["uuid"], is_admin=is_admin)
    if n == -1:
        raise HTTPException(403, "forbidden")
    if n == 0:
        raise HTTPException(404, "note not found")
    return {"deleted": note_uuid}


# ─── MEETINGS CRUD ───

class MeetingCreateReq(BaseModel):
    contact_uuid: Optional[str] = None
    company: Optional[str] = None
    person: Optional[str] = None
    meeting_date: Optional[str] = None
    location: Optional[str] = None
    city: Optional[str] = None
    notes: Optional[str] = None
    outcome: Optional[str] = None
    is_shared: Optional[bool] = False


@app.get("/api/meetings")
def meetings_list(contact_uuid: Optional[str] = None, user=Depends(current_user)):
    with db.db() as c:
        return db.list_meetings(c, user, contact_uuid=contact_uuid)


@app.post("/api/meetings")
def meetings_create(req: MeetingCreateReq, user=Depends(current_user)):
    with db.db() as c:
        muuid = db.create_meeting(c, owner_uuid=user["uuid"], **req.dict(exclude_none=True))
    return {"uuid": muuid}


@app.patch("/api/meetings/{meeting_uuid}")
def meetings_update(meeting_uuid: str, body: dict, user=Depends(current_user)):
    is_admin = user["role"] in ("super_admin", "admin")
    with db.db() as c:
        n = db.update_meeting(c, meeting_uuid, body or {}, user["uuid"], is_admin=is_admin)
    if n == -1:
        raise HTTPException(403, "forbidden")
    if n == 0:
        raise HTTPException(404, "meeting not found")
    return {"updated": n}


@app.delete("/api/meetings/{meeting_uuid}")
def meetings_delete(meeting_uuid: str, user=Depends(current_user)):
    is_admin = user["role"] in ("super_admin", "admin")
    with db.db() as c:
        n = db.delete_meeting(c, meeting_uuid, user["uuid"], is_admin=is_admin)
    if n == -1:
        raise HTTPException(403, "forbidden")
    if n == 0:
        raise HTTPException(404, "meeting not found")
    return {"deleted": meeting_uuid}


# ─── EVENTS / AUDIT ───

@app.get("/api/events")
def events_list(limit: int = 100, user: Optional[str] = None, type: Optional[str] = None,
                requester=Depends(current_user)):
    with db.db() as c:
        return db.list_events(c, limit=limit, user_uuid=user, event_type=type, requester=requester)


# Serve frontend — static assets + SPA fallback
if os.path.isdir(FRONTEND_BUILD):
    from fastapi.responses import FileResponse
    app.mount("/_app", StaticFiles(directory=os.path.join(FRONTEND_BUILD, "_app")), name="assets")

    @app.get("/{path:path}")
    async def spa_fallback(path: str):
        file_path = os.path.join(FRONTEND_BUILD, path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(FRONTEND_BUILD, "index.html"))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
